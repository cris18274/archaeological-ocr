"""
ArcheoOCR – Interfaz de Escritorio (Tkinter)
Conecta con el backend FastAPI en http://localhost:8000
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import requests
import json
import os
import io
import subprocess
import sys
import time
from pathlib import Path

# ─── Colores y estilos ────────────────────────────────────────────────────────
BG_DARK      = "#0f1117"
BG_CARD      = "#1a1d27"
BG_INPUT     = "#252836"
ACCENT       = "#4f8ef7"
ACCENT_HOVER = "#3a7de0"
SUCCESS      = "#2ecc71"
WARNING      = "#f39c12"
ERROR_COL    = "#e74c3c"
TXT_PRIMARY  = "#e8eaf6"
TXT_SECONDARY= "#8892a4"
TXT_MUTED    = "#55627a"
BORDER       = "#2a2d3e"

BACKEND_URL  = "http://localhost:8000"
CUSTOM_MODEL = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "custom_models", "archaeo_rec_v1", "inference")
COLOR_STD    = "#4f8ef7"   # azul para estándar
COLOR_CUS    = "#f7a24f"   # naranja para custom
COLOR_AGREE  = "#2ecc71"   # verde para coincidencia

# ─── Utilidades ───────────────────────────────────────────────────────────────

def start_server_if_needed():
    """Intenta conectar; si falla, lanza el servidor en background."""
    try:
        r = requests.get(f"{BACKEND_URL}/docs", timeout=3)
        return True
    except Exception:
        return False


class ArcheoOCRApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ArcheoOCR – Extracción Inteligente de Tablas")
        self.geometry("1300x820")
        self.minsize(900, 600)
        self.configure(bg=BG_DARK)

        # Estado
        self.selected_file = tk.StringVar(value="")
        self.status_text   = tk.StringVar(value="Listo. Selecciona un archivo para comenzar.")
        self.progress_val  = tk.DoubleVar(value=0.0)
        self.processing    = False
        self.last_result   = None
        self.engine_mode   = tk.StringVar(value="standard")  # standard | custom | compare

        self._apply_theme()
        self._build_ui()
        self._check_server_status()

    # ─── Tema ttk ──────────────────────────────────────────────────────────

    def _apply_theme(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        style.configure(".",
            background=BG_DARK, foreground=TXT_PRIMARY,
            fieldbackground=BG_INPUT, bordercolor=BORDER,
            troughcolor=BG_CARD, selectbackground=ACCENT,
            selectforeground=TXT_PRIMARY, font=("Segoe UI", 10)
        )
        style.configure("Treeview",
            background=BG_CARD, foreground=TXT_PRIMARY,
            fieldbackground=BG_CARD, rowheight=26,
            borderwidth=0, relief="flat"
        )
        style.configure("Treeview.Heading",
            background=BG_INPUT, foreground=ACCENT,
            relief="flat", font=("Segoe UI", 9, "bold")
        )
        style.map("Treeview",
            background=[("selected", ACCENT)],
            foreground=[("selected", TXT_PRIMARY)]
        )
        style.configure("Vertical.TScrollbar",
            background=BG_INPUT, troughcolor=BG_DARK,
            arrowcolor=TXT_SECONDARY, bordercolor=BORDER, width=10
        )
        style.configure("Horizontal.TScrollbar",
            background=BG_INPUT, troughcolor=BG_DARK,
            arrowcolor=TXT_SECONDARY, bordercolor=BORDER, width=10
        )
        style.configure("TProgressbar",
            background=ACCENT, troughcolor=BG_INPUT,
            bordercolor=BG_DARK, thickness=6
        )
        style.configure("TNotebook",
            background=BG_DARK, bordercolor=BORDER, tabmargins=0
        )
        style.configure("TNotebook.Tab",
            background=BG_CARD, foreground=TXT_SECONDARY,
            padding=[14, 6], font=("Segoe UI", 9)
        )
        style.map("TNotebook.Tab",
            background=[("selected", BG_INPUT)],
            foreground=[("selected", TXT_PRIMARY)]
        )

    # ─── Construcción UI ───────────────────────────────────────────────────

    def _build_ui(self):
        # ── Barra superior ──────────────────────────────────────────────
        topbar = tk.Frame(self, bg=BG_CARD, height=60)
        topbar.pack(fill="x", side="top")
        topbar.pack_propagate(False)

        tk.Label(topbar, text="⚗ ArcheoOCR",
                 bg=BG_CARD, fg=ACCENT,
                 font=("Segoe UI", 16, "bold")).pack(side="left", padx=20, pady=14)

        tk.Label(topbar, text="Extracción Inteligente de Tablas Arqueológicas",
                 bg=BG_CARD, fg=TXT_SECONDARY,
                 font=("Segoe UI", 10)).pack(side="left", pady=14)

        self._server_dot = tk.Label(topbar, text="●", bg=BG_CARD, fg=WARNING,
                                    font=("Segoe UI", 14))
        self._server_dot.pack(side="right", padx=8)
        self._server_lbl = tk.Label(topbar, text="Verificando servidor...",
                                    bg=BG_CARD, fg=TXT_SECONDARY,
                                    font=("Segoe UI", 9))
        self._server_lbl.pack(side="right")

        # ── Panel izquierdo ─────────────────────────────────────────────
        left = tk.Frame(self, bg=BG_DARK, width=280)
        left.pack(fill="y", side="left", padx=0)
        left.pack_propagate(False)

        self._build_left_panel(left)

        # ── Separador ───────────────────────────────────────────────────
        sep = tk.Frame(self, bg=BORDER, width=1)
        sep.pack(fill="y", side="left")

        # ── Area principal ──────────────────────────────────────────────
        main = tk.Frame(self, bg=BG_DARK)
        main.pack(fill="both", expand=True, side="left")

        self._build_main_area(main)

        # ── Barra de estado ─────────────────────────────────────────────
        statusbar = tk.Frame(self, bg=BG_CARD, height=32)
        statusbar.pack(fill="x", side="bottom")
        statusbar.pack_propagate(False)

        tk.Label(statusbar, textvariable=self.status_text,
                 bg=BG_CARD, fg=TXT_SECONDARY,
                 font=("Segoe UI", 9)).pack(side="left", padx=14)

        self._progress = ttk.Progressbar(statusbar, variable=self.progress_val,
                                          length=200, mode="indeterminate",
                                          style="TProgressbar")
        self._progress.pack(side="right", padx=14, pady=8)

    def _build_left_panel(self, parent):
        pad = dict(padx=14, pady=6)

        # Sección: Archivo
        self._section_label(parent, "📂  ARCHIVO DE ENTRADA")

        file_frame = tk.Frame(parent, bg=BG_CARD, bd=0,
                              highlightthickness=1, highlightbackground=BORDER)
        file_frame.pack(fill="x", padx=10, pady=4)

        self._file_lbl = tk.Label(file_frame,
                                   text="Sin archivo seleccionado",
                                   bg=BG_CARD, fg=TXT_MUTED,
                                   font=("Segoe UI", 9), wraplength=230,
                                   justify="left", anchor="w")
        self._file_lbl.pack(fill="x", padx=10, pady=8)

        self._btn(parent, "🗁  Seleccionar Archivo",
                  self._select_file, primary=False)

        # Sección: Motor OCR
        self._section_label(parent, "🤖  MOTOR OCR")
        engine_frame = tk.Frame(parent, bg=BG_DARK)
        engine_frame.pack(fill="x", padx=10, pady=2)
        for val, label, fg in [
            ("standard", "● Estándar (GPU)",    COLOR_STD),
            ("custom",   "● Custom archaeo_v1", COLOR_CUS),
            ("compare",  "⇄ Comparar ambos",    SUCCESS),
        ]:
            rb = tk.Radiobutton(engine_frame, text=label, variable=self.engine_mode,
                                value=val, bg=BG_DARK, fg=fg,
                                selectcolor=BG_INPUT, activebackground=BG_DARK,
                                activeforeground=fg, font=("Segoe UI", 9),
                                cursor="hand2")
            rb.pack(anchor="w", pady=1)

        # Sección: Proceso
        self._section_label(parent, "⚙  PROCESAMIENTO")

        self._process_btn = self._btn(parent, "▶  Iniciar Extracción",
                                       self._start_processing, primary=True)
        self._process_btn.config(state="disabled")

        # Sección: Exportar
        self._section_label(parent, "💾  EXPORTAR RESULTADOS")

        self._btn(parent, "📊  Exportar a Excel", self._export_excel, primary=False)
        self._btn(parent, "📄  Exportar a CSV",   self._export_csv,   primary=False)
        self._btn(parent, "📋  Exportar a JSON",  self._export_json,  primary=False)

        # Sección: Estadísticas
        self._section_label(parent, "📈  ESTADÍSTICAS")
        self._stats_frame = tk.Frame(parent, bg=BG_DARK)
        self._stats_frame.pack(fill="x", padx=10)
        self._update_stats(None)

    def _build_main_area(self, parent):
        # Notebook
        self._notebook = ttk.Notebook(parent, style="TNotebook")
        self._notebook.pack(fill="both", expand=True, padx=0, pady=0)

        # ── Tab: Tabla ───────────────────────────────────────────────────
        tab_table = tk.Frame(self._notebook, bg=BG_DARK)
        self._notebook.add(tab_table, text="  📊 Tabla Extraída  ")
        self._build_table_tab(tab_table)

        # ── Tab: JSON Raw ────────────────────────────────────────────────
        tab_json = tk.Frame(self._notebook, bg=BG_DARK)
        self._notebook.add(tab_json, text="  📋 JSON Completo  ")
        self._build_json_tab(tab_json)

        # ── Tab: Comparación ────────────────────────────────────────────
        tab_cmp = tk.Frame(self._notebook, bg=BG_DARK)
        self._notebook.add(tab_cmp, text="  ⇄ Comparación  ")
        self._build_compare_tab(tab_cmp)

        # ── Tab: Log ─────────────────────────────────────────────────────
        tab_log = tk.Frame(self._notebook, bg=BG_DARK)
        self._notebook.add(tab_log, text="  📝 Log  ")
        self._build_log_tab(tab_log)

    def _build_table_tab(self, parent):
        # Info badge row
        badge_row = tk.Frame(parent, bg=BG_DARK)
        badge_row.pack(fill="x", padx=10, pady=(8, 4))

        self._badge_rows = self._badge(badge_row, "Filas", "0")
        self._badge_cols = self._badge(badge_row, "Columnas", "0")
        self._badge_kws  = self._badge(badge_row, "Palabras clave", "0")

        # Search bar
        search_row = tk.Frame(parent, bg=BG_DARK)
        search_row.pack(fill="x", padx=10, pady=(0, 6))
        tk.Label(search_row, text="🔍 Buscar:", bg=BG_DARK, fg=TXT_SECONDARY,
                 font=("Segoe UI", 9)).pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._on_search)
        search_entry = tk.Entry(search_row, textvariable=self._search_var,
                                bg=BG_INPUT, fg=TXT_PRIMARY, insertbackground=TXT_PRIMARY,
                                relief="flat", font=("Segoe UI", 10), bd=6)
        search_entry.pack(side="left", fill="x", expand=True, padx=8)

        # Tabla con scrollbars
        table_frame = tk.Frame(parent, bg=BG_DARK)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self._tree = ttk.Treeview(table_frame, style="Treeview",
                                   show="headings", selectmode="extended")

        vsb = ttk.Scrollbar(table_frame, orient="vertical",
                             command=self._tree.yview, style="Vertical.TScrollbar")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal",
                             command=self._tree.xview, style="Horizontal.TScrollbar")
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(side="left", fill="both", expand=True)

        # Colores alternos de fila
        self._tree.tag_configure("odd",  background="#1e2132")
        self._tree.tag_configure("even", background=BG_CARD)
        self._tree.tag_configure("match", background="#2d3a1e", foreground="#a8e063")

        # Estado vacío
        self._empty_lbl = tk.Label(table_frame,
            text="Aquí aparecerá la tabla extraída\ndespués de procesar un documento.",
            bg=BG_DARK, fg=TXT_MUTED, font=("Segoe UI", 12),
            justify="center")
        self._empty_lbl.place(relx=0.5, rely=0.5, anchor="center")

    def _build_compare_tab(self, parent):
        """Tab con tablas STD y CUSTOM lado a lado."""
        top = tk.Frame(parent, bg=BG_DARK)
        top.pack(fill="x", padx=10, pady=6)

        # Badges de métricas de comparación
        self._cmp_badge_agree  = self._badge(top, "Coincidencia", "—")
        self._cmp_badge_filled_std = self._badge(top, "Celdas STD", "—")
        self._cmp_badge_filled_cus = self._badge(top, "Celdas Custom", "—")
        self._cmp_badge_agree.config(fg=SUCCESS)
        self._cmp_badge_filled_std.config(fg=COLOR_STD)
        self._cmp_badge_filled_cus.config(fg=COLOR_CUS)

        panes = tk.PanedWindow(parent, orient="horizontal",
                               bg=BORDER, sashwidth=4, sashrelief="flat")
        panes.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Panel izquierdo: estándar
        left_frame = tk.Frame(panes, bg=BG_DARK)
        tk.Label(left_frame, text="MODELO ESTÁNDAR",
                 bg=BG_DARK, fg=COLOR_STD,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=4, pady=4)
        self._cmp_tree_std = self._make_mini_tree(left_frame)
        panes.add(left_frame, stretch="always")

        # Panel derecho: custom
        right_frame = tk.Frame(panes, bg=BG_DARK)
        tk.Label(right_frame, text="MODELO CUSTOM (archaeo_rec_v1)",
                 bg=BG_DARK, fg=COLOR_CUS,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=4, pady=4)
        self._cmp_tree_cus = self._make_mini_tree(right_frame)
        panes.add(right_frame, stretch="always")

        self._cmp_empty = tk.Label(parent,
            text="Ejecuta una comparación (modo '⇄ Comparar ambos')\npara ver los resultados aquí.",
            bg=BG_DARK, fg=TXT_MUTED, font=("Segoe UI", 11), justify="center")
        self._cmp_empty.place(relx=0.5, rely=0.5, anchor="center")

    def _make_mini_tree(self, parent):
        frame = tk.Frame(parent, bg=BG_DARK)
        frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, style="Treeview", show="headings",
                             selectmode="browse")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        tree.tag_configure("odd",   background="#1e2132")
        tree.tag_configure("even",  background=BG_CARD)
        tree.tag_configure("diff",  background="#3a1e1e", foreground="#ff9090")
        tree.tag_configure("agree", background="#1e3a1e", foreground="#90ff90")
        return tree

    def _build_json_tab(self, parent):
        self._json_text = tk.Text(parent, bg=BG_CARD, fg="#a8d8ea",
                                   font=("Consolas", 9), relief="flat",
                                   wrap="none", state="disabled",
                                   insertbackground=TXT_PRIMARY)
        vsb = ttk.Scrollbar(parent, orient="vertical",
                             command=self._json_text.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal",
                             command=self._json_text.xview)
        self._json_text.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self._json_text.pack(fill="both", expand=True)

    def _build_log_tab(self, parent):
        self._log_text = tk.Text(parent, bg=BG_CARD, fg="#90ee90",
                                  font=("Consolas", 9), relief="flat",
                                  wrap="word", state="disabled",
                                  insertbackground=TXT_PRIMARY)
        vsb = ttk.Scrollbar(parent, orient="vertical",
                             command=self._log_text.yview)
        self._log_text.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._log_text.pack(fill="both", expand=True, padx=4, pady=4)

        self._log(f"[{self._ts()}] ArcheoOCR iniciado.\n")

    # ─── Widgets auxiliares ────────────────────────────────────────────────

    def _section_label(self, parent, text):
        tk.Label(parent, text=text,
                 bg=BG_DARK, fg=TXT_MUTED,
                 font=("Segoe UI", 8, "bold")).pack(
            anchor="w", padx=14, pady=(14, 2))

    def _btn(self, parent, text, cmd, primary=True):
        bg = ACCENT if primary else BG_INPUT
        fg = TXT_PRIMARY
        btn = tk.Button(parent, text=text, command=cmd,
                        bg=bg, fg=fg, relief="flat",
                        font=("Segoe UI", 9, "bold" if primary else "normal"),
                        cursor="hand2", anchor="w",
                        padx=14, pady=8, activebackground=ACCENT_HOVER,
                        activeforeground=TXT_PRIMARY)
        btn.pack(fill="x", padx=10, pady=2)

        def on_enter(e): btn.config(bg=ACCENT_HOVER if primary else BG_CARD)
        def on_leave(e): btn.config(bg=bg)
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    def _badge(self, parent, label, value):
        frame = tk.Frame(parent, bg=BG_INPUT, padx=12, pady=6)
        frame.pack(side="left", padx=(0, 8))
        lbl = tk.Label(frame, text=value,
                       bg=BG_INPUT, fg=ACCENT,
                       font=("Segoe UI", 16, "bold"))
        lbl.pack()
        tk.Label(frame, text=label,
                 bg=BG_INPUT, fg=TXT_SECONDARY,
                 font=("Segoe UI", 8)).pack()
        return lbl  # retornamos el Label del valor para actualizar

    def _update_stats(self, data):
        for w in self._stats_frame.winfo_children():
            w.destroy()

        if data is None:
            items = [("Filas", "—"), ("Columnas", "—"), ("Tablas", "—")]
        else:
            rows = len(data.get("rows", []))
            cols = len(data.get("header", []))
            items = [("Filas", str(rows)), ("Columnas", str(cols)),
                     ("KWs", str(len(data.get("keywords", []))))]

        for label, val in items:
            row = tk.Frame(self._stats_frame, bg=BG_DARK)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=label, bg=BG_DARK, fg=TXT_SECONDARY,
                     font=("Segoe UI", 9), width=10, anchor="w").pack(side="left")
            tk.Label(row, text=val, bg=BG_DARK, fg=TXT_PRIMARY,
                     font=("Segoe UI", 9, "bold"), anchor="e").pack(side="right")

    # ─── Lógica principal ──────────────────────────────────────────────────

    def _check_server_status(self):
        def check():
            try:
                r = requests.get(f"{BACKEND_URL}/docs", timeout=3)
                ok = r.status_code == 200
            except Exception:
                ok = False
            self.after(0, lambda: self._set_server_status(ok))

        threading.Thread(target=check, daemon=True).start()
        self.after(10000, self._check_server_status)   # re-chequear cada 10s

    def _set_server_status(self, ok):
        if ok:
            self._server_dot.config(fg=SUCCESS)
            self._server_lbl.config(text=f"Servidor OK  ({BACKEND_URL})",
                                    fg=SUCCESS)
        else:
            self._server_dot.config(fg=ERROR_COL)
            self._server_lbl.config(text="Servidor no disponible",
                                    fg=ERROR_COL)

    def _select_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar documento",
            filetypes=[
                ("Documentos soportados", "*.pdf *.png *.jpg *.jpeg"),
                ("PDF", "*.pdf"),
                ("Imágenes", "*.png *.jpg *.jpeg"),
                ("Todos", "*.*"),
            ]
        )
        if path:
            self.selected_file.set(path)
            name = os.path.basename(path)
            size = os.path.getsize(path) / 1024
            self._file_lbl.config(
                text=f"{name}\n{size:.1f} KB", fg=TXT_PRIMARY)
            self._process_btn.config(state="normal")
            self._log(f"[{self._ts()}] Archivo seleccionado: {path}\n")
            self.status_text.set(f"Archivo listo: {name}")

    def _start_processing(self):
        if self.processing:
            return
        path = self.selected_file.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Selecciona un archivo válido primero.")
            return

        mode = self.engine_mode.get()
        self.processing = True
        self._process_btn.config(state="disabled", text="⏳  Procesando...")
        self._progress.start(12)
        self._log(f"\n[{self._ts()}] Modo: {mode.upper()} | Archivo: {os.path.basename(path)}\n")

        if mode == "compare":
            self.status_text.set("Comparando ambos motores... (×2 tiempo)")
            threading.Thread(target=self._run_compare, args=(path,), daemon=True).start()
        else:
            self.status_text.set(f"Procesando con motor {mode}... (1–2 min GPU)")
            threading.Thread(target=self._run_extraction,
                             args=(path, mode), daemon=True).start()

    def _run_extraction(self, path, engine="standard"):
        t0 = time.time()
        try:
            with open(path, "rb") as f:
                mime = "application/pdf" if path.lower().endswith(".pdf") else "image/jpeg"
                # Pasamos el engine como query param
                resp = requests.post(
                    f"{BACKEND_URL}/upload",
                    files={"file": (os.path.basename(path), f, mime)},
                    params={"engine": engine},
                    timeout=300
                )

            elapsed = time.time() - t0

            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:300]}")

            data = resp.json()
            self.last_result = data
            self.after(0, lambda d=data, e=elapsed: self._on_result(d, e, engine=engine))

        except Exception as e:
            self.after(0, lambda err=e: self._on_error(str(err)))

    def _run_compare(self, path):
        """Llama al backend dos veces y muestra comparación."""
        t0 = time.time()
        mime = "application/pdf" if path.lower().endswith(".pdf") else "image/jpeg"
        results = {}
        for engine in ("standard", "custom"):
            self.after(0, lambda e=engine:
                self.status_text.set(f"Procesando motor {e.upper()}..."))
            try:
                with open(path, "rb") as f:
                    resp = requests.post(
                        f"{BACKEND_URL}/upload",
                        files={"file": (os.path.basename(path), f, mime)},
                        params={"engine": engine},
                        timeout=300
                    )
                if resp.status_code == 200:
                    results[engine] = resp.json()
                else:
                    results[engine] = None
                    self._log(f"[{self._ts()}] [{engine}] Error HTTP {resp.status_code}\n")
            except Exception as e:
                results[engine] = None
                self._log(f"[{self._ts()}] [{engine}] Excepción: {e}\n")

        elapsed = time.time() - t0
        self.after(0, lambda: self._on_compare_result(
            results.get("standard"), results.get("custom"), elapsed))

    def _on_result(self, data, elapsed, engine="standard"):
        self._progress.stop()
        self.processing = False
        self._process_btn.config(state="normal", text="▶  Iniciar Extracción")

        table_content = self._extract_first_table(data)
        if not table_content:
            self.status_text.set("✓ Proceso completado — No se detectaron tablas.")
            self._log(f"[{self._ts()}] Sin tablas detectadas.\n")
            messagebox.showinfo("Resultado",
                "El procesamiento terminó pero no se encontraron tablas en el documento.")
            return

        header = table_content.get("header", [])
        rows   = table_content.get("rows",   [])
        kws    = table_content.get("keywords", [])

        color = COLOR_STD if engine == "standard" else COLOR_CUS
        self._badge_rows.config(text=str(len(rows)), fg=color)
        self._badge_cols.config(text=str(len(header)), fg=color)
        self._badge_kws.config( text=str(len(kws)))

        self._populate_table(header, rows)
        self._update_stats(table_content)
        self._set_json(json.dumps(data, ensure_ascii=False, indent=2))
        self._notebook.select(0)
        self._empty_lbl.place_forget()

        tag = "[STD]" if engine == "standard" else "[CUSTOM]"
        msg = (f"✓ {tag} {elapsed:.1f}s — "
               f"{len(rows)} filas × {len(header)} columnas")
        self.status_text.set(msg)
        self._log(f"[{self._ts()}] {msg}\n")
        if kws:
            self._log(f"[{self._ts()}] KWs: {', '.join(kws[:10])}\n")

    def _on_compare_result(self, std_data, cus_data, elapsed):
        self._progress.stop()
        self.processing = False
        self._process_btn.config(state="normal", text="▶  Iniciar Extracción")

        std_content = self._extract_first_table(std_data) if std_data else None
        cus_content = self._extract_first_table(cus_data) if cus_data else None

        if not std_content and not cus_content:
            messagebox.showwarning("Sin resultados",
                "Ningún motor pudo extraer tablas del documento.")
            return

        # Poblar tab principal con el resultado estándar
        if std_content:
            self._badge_rows.config(text=str(len(std_content.get("rows", []))),
                                    fg=COLOR_STD)
            self._badge_cols.config(text=str(len(std_content.get("header", []))),
                                    fg=COLOR_STD)
            self._populate_table(std_content.get("header", []),
                                  std_content.get("rows", []))
            self._empty_lbl.place_forget()

        # Poblar tab de comparación
        self._populate_compare_tab(std_content, cus_content)
        self._notebook.select(2)   # ir a tab Comparación
        self._cmp_empty.place_forget()

        # JSON con ambos resultados
        combined = {"standard": std_data, "custom": cus_data}
        self._set_json(json.dumps(combined, ensure_ascii=False, indent=2))
        if std_data:
            self.last_result = std_data

        msg = f"✓ Comparación en {elapsed:.1f}s"
        self.status_text.set(msg)
        self._log(f"[{self._ts()}] {msg}\n")

    def _extract_first_table(self, data):
        if not data:
            return None
        for page in data.get("pages", []):
            for region in page.get("regions", []):
                if region.get("type") == "table" and region.get("content"):
                    return region["content"]
        return None

    def _populate_compare_tab(self, std_content, cus_content):
        import difflib
        std_h   = std_content.get("header", []) if std_content else []
        std_rows = std_content.get("rows", [])  if std_content else []
        cus_h   = cus_content.get("header", []) if cus_content else []
        cus_rows = cus_content.get("rows", [])  if cus_content else []

        max_cols = max(len(std_h), len(cus_h), 1)
        all_cols = [str(i) for i in range(max_cols)]

        def fill_tree(tree, header, rows, ref_rows):
            tree.delete(*tree.get_children())
            cols = [str(i) for i in range(max_cols)]
            tree.configure(columns=cols)
            for i, h in enumerate(header):
                tree.heading(i, text=str(h) if h else f"C{i+1}", anchor="w")
                tree.column(i, width=140 if i == 0 else 50, minwidth=35)
            nrow = max(len(rows), len(ref_rows))
            for r_i in range(nrow):
                row = rows[r_i] if r_i < len(rows) else []
                ref = ref_rows[r_i] if r_i < len(ref_rows) else []
                vals = [str(c) if c else "" for c in row]
                while len(vals) < max_cols:
                    vals.append("")
                ref_vals = [str(c) if c else "" for c in ref]
                # Colorear filas que difieren en col 0
                s0 = vals[0].strip() if vals else ""
                r0 = ref_vals[0].strip() if ref_vals else ""
                ratio = difflib.SequenceMatcher(None, s0, r0).ratio()
                if ratio >= 0.85:
                    tag = "agree"
                elif ratio < 0.4:
                    tag = "diff"
                else:
                    tag = "odd" if r_i % 2 else "even"
                tree.insert("", "end", iid=str(r_i), values=vals, tags=(tag,))

        fill_tree(self._cmp_tree_std, std_h, std_rows, cus_rows)
        fill_tree(self._cmp_tree_cus, cus_h, cus_rows, std_rows)

        # Métricas rápidas
        total = max(len(std_rows), len(cus_rows))
        matches = sum(
            1 for i in range(min(len(std_rows), len(cus_rows)))
            if (std_rows[i][0] if std_rows[i] else "") ==
               (cus_rows[i][0] if cus_rows[i] else "")
        )
        pct = round(matches / max(total, 1) * 100)
        s_filled = sum(1 for r in std_rows for c in r if c and str(c).strip())
        c_filled = sum(1 for r in cus_rows for c in r if c and str(c).strip())

        self._cmp_badge_agree.config(text=f"{pct}%")
        self._cmp_badge_filled_std.config(text=str(s_filled))
        self._cmp_badge_filled_cus.config(text=str(c_filled))

        self._log(f"[{self._ts()}] Comparación: {pct}% coincidencia | "
                  f"STD {s_filled} celdas | CUS {c_filled} celdas\n")

    def _on_error(self, msg):
        self._progress.stop()
        self.processing = False
        self._process_btn.config(state="normal", text="▶  Iniciar Extracción")
        self.status_text.set(f"❌ Error: {msg[:80]}")
        self._log(f"[{self._ts()}] ERROR: {msg}\n")
        messagebox.showerror("Error de procesamiento", msg)

    # ─── Tabla ─────────────────────────────────────────────────────────────

    def _populate_table(self, header, rows):
        tree = self._tree
        tree.delete(*tree.get_children())

        # Columnas
        columns = [str(i) for i in range(len(header))]
        tree.configure(columns=columns)

        # Cabeceras
        for i, h in enumerate(header):
            text = str(h) if h else f"Col {i+1}"
            tree.heading(i, text=text, anchor="w")
            # Ancho auto: más ancho para la primera columna
            w = 160 if i == 0 else 55
            tree.column(i, width=w, minwidth=40, anchor="w")

        # Filas
        self._all_rows = (header, rows)
        for idx, row in enumerate(rows):
            vals = [str(c) if c is not None else "" for c in row]
            # Rellenar si la fila es más corta que el header
            while len(vals) < len(header):
                vals.append("")
            tag = "odd" if idx % 2 else "even"
            tree.insert("", "end", iid=str(idx), values=vals, tags=(tag,))

    def _on_search(self, *_):
        if not self._all_rows:
            return
        header, rows = self._all_rows
        q = self._search_var.get().strip().lower()
        tree = self._tree
        tree.delete(*tree.get_children())

        for idx, row in enumerate(rows):
            vals = [str(c) if c is not None else "" for c in row]
            while len(vals) < len(header):
                vals.append("")
            match = q and any(q in v.lower() for v in vals)
            tag = "match" if match else ("odd" if idx % 2 else "even")
            tree.insert("", "end", iid=str(idx), values=vals, tags=(tag,))

        # Scroll al primer match
        if q:
            for child in tree.get_children():
                if "match" in tree.item(child, "tags"):
                    tree.see(child)
                    break

    # ─── Exportar ──────────────────────────────────────────────────────────

    def _get_table_data(self):
        if not self.last_result:
            messagebox.showwarning("Sin datos", "Primero procesa un documento.")
            return None, None
        pages = self.last_result.get("pages", [])
        for page in pages:
            for region in page.get("regions", []):
                if region.get("type") == "table" and region.get("content"):
                    c = region["content"]
                    return c.get("header", []), c.get("rows", [])
        messagebox.showwarning("Sin datos", "No se encontró ninguna tabla en el resultado.")
        return None, None

    def _export_excel(self):
        header, rows = self._get_table_data()
        if header is None:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar como Excel"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tabla OCR"

            # Header con estilo
            for col_i, h in enumerate(header, 1):
                cell = ws.cell(row=1, column=col_i, value=str(h) if h else f"Col{col_i}")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1a3a6e")
                cell.alignment = Alignment(horizontal="center")

            # Datos
            for row_i, row in enumerate(rows, 2):
                for col_i, val in enumerate(row, 1):
                    ws.cell(row=row_i, column=col_i, value=str(val) if val else "")

            # Autofit columnas
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

            wb.save(path)
            self._log(f"[{self._ts()}] Excel guardado: {path}\n")
            messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")
        except ImportError:
            messagebox.showerror("Error", "Instala openpyxl: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _export_csv(self):
        import csv
        header, rows = self._get_table_data()
        if header is None:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            title="Guardar como CSV"
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(header)
                w.writerows(rows)
            self._log(f"[{self._ts()}] CSV guardado: {path}\n")
            messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _export_json(self):
        if not self.last_result:
            messagebox.showwarning("Sin datos", "Primero procesa un documento.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            title="Guardar como JSON"
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.last_result, f, ensure_ascii=False, indent=2)
            self._log(f"[{self._ts()}] JSON guardado: {path}\n")
            messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ─── Helpers ──────────────────────────────────────────────────────────

    def _log(self, msg):
        self._log_text.config(state="normal")
        self._log_text.insert("end", msg)
        self._log_text.see("end")
        self._log_text.config(state="disabled")

    def _set_json(self, text):
        self._json_text.config(state="normal")
        self._json_text.delete("1.0", "end")
        self._json_text.insert("1.0", text)
        self._json_text.config(state="disabled")

    def _ts(self):
        return time.strftime("%H:%M:%S")

    # init para _all_rows
    _all_rows = None


# ─── Splash de inicio ─────────────────────────────────────────────────────────

def show_splash(root):
    splash = tk.Toplevel(root)
    splash.overrideredirect(True)
    splash.configure(bg=BG_CARD)

    w, h = 420, 200
    sw = splash.winfo_screenwidth()
    sh = splash.winfo_screenheight()
    splash.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    tk.Label(splash, text="⚗ ArcheoOCR", bg=BG_CARD, fg=ACCENT,
             font=("Segoe UI", 24, "bold")).pack(pady=(30, 4))
    tk.Label(splash, text="Extracción Inteligente de Tablas Arqueológicas",
             bg=BG_CARD, fg=TXT_SECONDARY,
             font=("Segoe UI", 10)).pack()

    prog = ttk.Progressbar(splash, mode="indeterminate", length=340)
    prog.pack(pady=24)
    prog.start(10)

    tk.Label(splash, text="Iniciando...", bg=BG_CARD, fg=TXT_MUTED,
             font=("Segoe UI", 9)).pack()

    splash.update()
    return splash


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = ArcheoOCRApp()
    root.withdraw()  # ocultar mientras carga

    splash = show_splash(root)
    root.after(2000, lambda: (splash.destroy(), root.deiconify()))

    root.mainloop()
